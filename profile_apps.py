#!/usr/bin/env python2.7

"""
    Profiles the apps whose source code is stored in the given folder.
"""

# --- imports
from collections import namedtuple
import os
import platform
import shutil
import subprocess
import sys
import tempfile
import time

from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.styles import colors, PatternFill, Fill

# --- constants
DELIMITER_LINE = "### App Intent Stats"
OUTPUT_FILE = "profiling_results.xlsx"
INTERMEDIATE_RESULTS_FOLDER = tempfile.mkdtemp()
HEADER = ["Intents", "Explicit Intents", "Implicit Intents", "ICC Links",
          "Explicit ICC Links", "Implicit ICC Links", "'startActivity' calls", "'startService' calls",
          "Elapsed Time (seconds)", "Explicit Intents (%)", "Explicit ICC Links (%)"]

# types
exec_result = namedtuple("ExecResult", "result_file elapsed_time")


# --- functions
def get_results_from_file(filepath):

    num = lambda f: int(f.readline().strip().split(":")[1].strip().rstrip())

    with open(filepath) as f:

        line = f.readline()
        while line:

            line = line.strip()

            if line == DELIMITER_LINE:

                partial_result = (num(f), num(f), num(f), num(f), num(f), num(f), num(f), num(f))

                return partial_result + (((partial_result[1] / (partial_result[0] * 1.0)) * 100) if partial_result[0] else 0,
                                         ((partial_result[4] / (partial_result[3] * 1.0)) * 100) if partial_result[3] else 0)

            line = f.readline()

    return (0, 0, 0, 0, 0, 0, 0, 0, 0, 0)


def fill_worksheet(worksheet, app_name, result_file, elapsed_time):

    # set the columns width
    for i, name in enumerate(HEADER):

        worksheet.column_dimensions[get_column_letter(i+1)].width = len(name)

    worksheet.append([""])
    worksheet.append([app_name])

    results = get_results_from_file(result_file)

    worksheet.append(results[:-2] + (elapsed_time,) + results[-2:])


def fill_final_results_worksheet(worksheet, results):

    worksheet.append(HEADER)

    # set the columns width
    for i, name in enumerate(HEADER):

        worksheet.column_dimensions[get_column_letter(i+1)].width = len(name)

    final_results = (0, 0, 0, 0, 0, 0, 0, 0, 0)

    for result in results.values():

        retrieved = get_results_from_file(result.result_file) + (result.elapsed_time,)

        final_results = tuple(final_results[i] + retrieved[i] for i in xrange(9))

    final_results = list(final_results)

    final_results.append(((final_results[1] / (final_results[0] * 1.0)) * 100) if final_results[0] else 0)
    final_results.append(((final_results[4] / (final_results[3] * 1.0)) * 100) if final_results[3] else 0)

    worksheet.append(final_results)


def main():

    APPS_FOLDER = sys.argv[1]
    results = dict()

    # run the pbicc tool on each of the apps
    for app in os.listdir(APPS_FOLDER):

        full_app_path = os.path.join(APPS_FOLDER, app)
        result_file = os.path.join(INTERMEDIATE_RESULTS_FOLDER, app)

        with open(result_file, 'w') as result_file_handler:

            print "Profiling app '%s'" % app

            cmd = ["./run.sh", full_app_path]

            starting_time = time.time()
            subprocess.Popen(cmd, stderr=subprocess.STDOUT, stdout=result_file_handler).wait()
            elapsed_time = time.time() - starting_time

            results[app] = exec_result(result_file, elapsed_time)

            print "done (in '%s' seconds)" % elapsed_time

    # creating the workbook
    wb = Workbook()

    ws = wb.active
    ws.title = "Final Results"

    fill_final_results_worksheet(ws, results)

    detailed_ws = wb.create_sheet(title="Detailed Results")
    detailed_ws.append(HEADER)

    # parse each of the apps
    for app, result_info in results.items():

        fill_worksheet(detailed_ws, app, result_info.result_file, result_info.elapsed_time)

    wb.save(OUTPUT_FILE)

    shutil.rmtree(INTERMEDIATE_RESULTS_FOLDER)

if __name__ == "__main__":

    main()